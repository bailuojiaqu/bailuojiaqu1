from app.models import  CustomName,CustomMonth,FormTotal
from common.mysql import session,engine
from common.redis import connect_redis
import openpyxl
import os


def get_custom_name():
    data = session.query(FormTotal.id,FormTotal.customname,FormTotal.year,FormTotal.date,FormTotal.size,FormTotal.num,FormTotal.price,FormTotal.price_total).all()
    engine.dispose()
    return data

def create_excel():
    dic_lst=[]
    lst=[]
    num_list=[]
    price_lst=[]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sheet"
    row = 2
    # 效果是在的第二行
    title = ['日期','尺寸规格','条数','单价','金额']
    for i in title:
        sheet.cell(1, title.index(i)+1, i)
    table = get_custom_name()
    for item in table:
        rows = {'month':item.date,'day':item.size,'num':item.num,'price':item.price,'price_total':item.price_total}
        dic_lst.append(rows)
        num_list.append(rows['num'])
        price_lst.append(rows['price_total'])
        for value in rows.values():
            lst.append(value)
        for i in range(len(lst)):
            sheet.cell(row, i+1, lst[i])
        row+=1
        lst.clear()
    sheet.cell(row, 1, '总计')
    sheet.cell(row, 3, sum(num_list))
    sheet.cell(row, 5, sum(price_lst))
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    workbook.save(desktop + "/table.xlsx")  
    return 0